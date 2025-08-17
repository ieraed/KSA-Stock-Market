"""
Portfolio Template Generator
Creates Excel templates for multi-market portfolio uploads
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
import numpy as np

def create_portfolio_template():
    """Create comprehensive portfolio template with sample data"""
    
    # Sample data for both Saudi and US markets
    sample_data = {
        'Symbol': [
            # US Stocks
            'AAPL', 'GOOGL', 'MSFT', 'TSLA', 'AMZN', 'META', 'NVDA', 'NFLX',
            # Saudi Stocks
            '2222.SR', '1010.SR', '2380.SR', '1120.SR', '2030.SR', '1180.SR', '2010.SR', '4030.SR'
        ],
        'Market': [
            # US Markets
            'US', 'US', 'US', 'US', 'US', 'US', 'US', 'US',
            # Saudi Market
            'Saudi', 'Saudi', 'Saudi', 'Saudi', 'Saudi', 'Saudi', 'Saudi', 'Saudi'
        ],
        'Exchange': [
            # US Exchanges
            'NASDAQ', 'NASDAQ', 'NASDAQ', 'NASDAQ', 'NASDAQ', 'NASDAQ', 'NASDAQ', 'NASDAQ',
            # Saudi Exchange
            'Tadawul', 'Tadawul', 'Tadawul', 'Tadawul', 'Tadawul', 'Tadawul', 'Tadawul', 'Tadawul'
        ],
        'Company_Name': [
            # US Companies
            'Apple Inc.', 'Alphabet Inc.', 'Microsoft Corp', 'Tesla Inc.', 'Amazon.com Inc', 
            'Meta Platforms Inc', 'NVIDIA Corp', 'Netflix Inc',
            # Saudi Companies
            'Saudi Basic Industries Corp (SABIC)', 'Saudi National Bank', 'Petrochemical Industries Co', 
            'Al Rajhi Bank', 'Saudi Aramco', 'National Commercial Bank', 'Saudi Electricity Co', 
            'National Medical Care Company'
        ],
        'Shares': [
            # Sample quantities
            100, 25, 150, 50, 30, 75, 40, 20,
            200, 500, 100, 300, 1000, 250, 150, 80
        ],
        'Purchase_Price': [
            # US stock prices (USD)
            175.50, 2750.00, 380.25, 850.00, 3200.00, 320.75, 450.00, 580.00,
            # Saudi stock prices (SAR)
            120.50, 85.75, 45.20, 40.80, 35.25, 55.60, 28.90, 95.40
        ],
        'Purchase_Date': [
            # Sample purchase dates
            '2024-01-15', '2024-01-20', '2024-02-01', '2024-02-15', '2024-03-01',
            '2024-03-15', '2024-04-01', '2024-04-15',
            '2024-01-10', '2024-02-05', '2024-02-20', '2024-03-10', '2024-03-25',
            '2024-04-05', '2024-04-20', '2024-05-01'
        ],
        'Currency': [
            # Currencies
            'USD', 'USD', 'USD', 'USD', 'USD', 'USD', 'USD', 'USD',
            'SAR', 'SAR', 'SAR', 'SAR', 'SAR', 'SAR', 'SAR', 'SAR'
        ],
        'Sector': [
            # US sectors
            'Technology', 'Technology', 'Technology', 'Automotive', 'E-commerce',
            'Technology', 'Technology', 'Entertainment',
            # Saudi sectors
            'Petrochemicals', 'Banking', 'Petrochemicals', 'Banking', 'Energy',
            'Banking', 'Utilities', 'Healthcare'
        ],
        'Industry': [
            # US industries
            'Consumer Electronics', 'Internet Services', 'Software', 'Electric Vehicles', 'E-commerce',
            'Social Media', 'Semiconductors', 'Streaming',
            # Saudi industries
            'Basic Materials', 'Commercial Banking', 'Chemicals', 'Commercial Banking', 'Oil & Gas',
            'Commercial Banking', 'Electric Utilities', 'Medical Services'
        ],
        'Target_Price': [
            # Sample target prices
            200.00, 3000.00, 420.00, 950.00, 3500.00, 380.00, 520.00, 650.00,
            135.00, 95.00, 52.00, 48.00, 42.00, 65.00, 35.00, 110.00
        ],
        'Stop_Loss': [
            # Sample stop loss prices
            150.00, 2400.00, 340.00, 750.00, 2800.00, 280.00, 380.00, 520.00,
            105.00, 75.00, 38.00, 35.00, 30.00, 48.00, 25.00, 85.00
        ],
        'Notes': [
            # Sample notes
            'Core holding - dividend growth', 'Long-term AI play', 'Cloud computing leader', 'EV leader position',
            'E-commerce dominance', 'Metaverse investment', 'AI chip leader', 'Content streaming',
            'Saudi basic materials leader', 'Largest Saudi bank', 'Petrochemical growth', 'Digital banking leader',
            'National oil company', 'Commercial banking', 'Utility monopoly', 'Healthcare growth'
        ]
    }
    
    # Create DataFrame
    df = pd.DataFrame(sample_data)
    
    # Save to Excel with formatting
    filename = f"portfolio_template_multi_market_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write main portfolio sheet
        df.to_excel(writer, sheet_name='Portfolio', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Portfolio']
        
        # Define colors
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        us_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        saudi_fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')
        
        # Define fonts
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        regular_font = Font(name='Arial', size=10)
        
        # Define alignment
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        # Format data rows
        for row_num in range(2, len(df) + 2):
            market_value = worksheet.cell(row=row_num, column=2).value  # Market column
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = regular_font
                cell.border = thin_border
                
                # Color code by market
                if market_value == 'US':
                    cell.fill = us_fill
                elif market_value == 'Saudi':
                    cell.fill = saudi_fill
                
                # Center align specific columns
                if col_num in [2, 3, 8]:  # Market, Exchange, Currency columns
                    cell.alignment = center_align
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create instructions sheet
        instructions_data = {
            'Field': [
                'Symbol', 'Market', 'Exchange', 'Company_Name', 'Shares', 'Purchase_Price',
                'Purchase_Date', 'Currency', 'Sector', 'Industry', 'Target_Price', 'Stop_Loss', 'Notes'
            ],
            'Description': [
                'Stock ticker symbol (e.g., AAPL for US, 2222.SR for Saudi)',
                'Market location: US or Saudi',
                'Stock exchange: NASDAQ, NYSE, AMEX for US; Tadawul for Saudi',
                'Full company name',
                'Number of shares owned',
                'Price per share when purchased',
                'Date of purchase (YYYY-MM-DD format)',
                'Currency: USD for US stocks, SAR for Saudi stocks',
                'Business sector (e.g., Technology, Banking, Energy)',
                'Specific industry within sector',
                'Your target selling price (optional)',
                'Your stop-loss price (optional)',
                'Any additional notes about the position'
            ],
            'Required': [
                'Yes', 'Yes', 'Yes', 'No', 'Yes', 'Yes',
                'Yes', 'Yes', 'No', 'No', 'No', 'No', 'No'
            ],
            'Examples': [
                'AAPL, GOOGL, 2222.SR, 1010.SR',
                'US, Saudi',
                'NASDAQ, NYSE, Tadawul',
                'Apple Inc., Saudi National Bank',
                '100, 500, 1000',
                '175.50, 85.75',
                '2024-01-15, 2024-02-01',
                'USD, SAR',
                'Technology, Banking, Energy',
                'Consumer Electronics, Commercial Banking',
                '200.00, 95.00',
                '150.00, 75.00',
                'Long-term hold, Dividend stock'
            ]
        }
        
        instructions_df = pd.DataFrame(instructions_data)
        instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
        
        # Format instructions sheet
        inst_worksheet = writer.sheets['Instructions']
        
        # Format headers
        for col_num, column_title in enumerate(instructions_df.columns, 1):
            cell = inst_worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        # Format data rows
        for row_num in range(2, len(instructions_df) + 2):
            for col_num in range(1, len(instructions_df.columns) + 1):
                cell = inst_worksheet.cell(row=row_num, column=col_num)
                cell.font = regular_font
                cell.border = thin_border
                
                # Highlight required fields
                if col_num == 3:  # Required column
                    required_value = cell.value
                    if required_value == 'Yes':
                        cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        
        # Auto-adjust column widths for instructions
        for column in inst_worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 60)
            inst_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create market info sheet
        market_info = {
            'Market': ['US Markets', 'US Markets', 'US Markets', 'Saudi Market'],
            'Exchange': ['NASDAQ', 'NYSE', 'AMEX', 'Tadawul'],
            'Currency': ['USD', 'USD', 'USD', 'SAR'],
            'Trading_Hours_Local': ['9:30 AM - 4:00 PM EST', '9:30 AM - 4:00 PM EST', 
                                   '9:30 AM - 4:00 PM EST', '10:00 AM - 3:00 PM AST'],
            'Trading_Days': ['Monday - Friday', 'Monday - Friday', 'Monday - Friday', 'Sunday - Thursday'],
            'Symbol_Format': ['AAPL, GOOGL, MSFT', 'IBM, JPM, GE', 'Various', '2222.SR, 1010.SR'],
            'Examples': ['Apple, Google, Microsoft', 'IBM, JPMorgan, GE', 'Smaller companies', 'SABIC, SNB']
        }
        
        market_df = pd.DataFrame(market_info)
        market_df.to_excel(writer, sheet_name='Market_Info', index=False)
        
        # Format market info sheet
        market_worksheet = writer.sheets['Market_Info']
        
        # Format headers
        for col_num, column_title in enumerate(market_df.columns, 1):
            cell = market_worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        # Format data rows
        for row_num in range(2, len(market_df) + 2):
            market_value = market_worksheet.cell(row=row_num, column=1).value
            
            for col_num in range(1, len(market_df.columns) + 1):
                cell = market_worksheet.cell(row=row_num, column=col_num)
                cell.font = regular_font
                cell.border = thin_border
                
                # Color code by market
                if 'US' in str(market_value):
                    cell.fill = us_fill
                elif 'Saudi' in str(market_value):
                    cell.fill = saudi_fill
        
        # Auto-adjust column widths for market info
        for column in market_worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 40)
            market_worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✅ Portfolio template created: {filename}")
    print("\n📊 Template includes:")
    print("  - Portfolio sheet with sample data (8 US + 8 Saudi stocks)")
    print("  - Instructions sheet with field descriptions")
    print("  - Market_Info sheet with trading details")
    print("\n🎨 Features:")
    print("  - Color-coded by market (Blue: US, Green: Saudi)")
    print("  - Required fields highlighted")
    print("  - Auto-sized columns")
    print("  - Professional formatting")
    
    return filename

if __name__ == "__main__":
    create_portfolio_template()
