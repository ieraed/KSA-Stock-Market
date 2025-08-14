#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Create updated portfolio template with corrected cost prices
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_corrected_portfolio_template():
    """Create Excel template with corrected portfolio data"""
    
    # Corrected portfolio data
    corrected_portfolio = [
        {"Company": "ALINMA", "Symbol": "1150", "Owned_Qty": 3722, "Cost": 27.02, "Custodian": "Al Inma Capital"},
        {"Company": "ACIG", "Symbol": "8150", "Owned_Qty": 1500, "Cost": 15.14, "Custodian": "Al Inma Capital"},
        {"Company": "NADEC", "Symbol": "2190", "Owned_Qty": 2967, "Cost": 22.28, "Custodian": "Al Inma Capital"},
        {"Company": "SUMOU", "Symbol": "4323", "Owned_Qty": 1300, "Cost": 42.86, "Custodian": "Al Inma Capital"},
        {"Company": "ALBILAD", "Symbol": "1140", "Owned_Qty": 1500, "Cost": 27.31, "Custodian": "Al Inma Capital"},
        {"Company": "SAUDI ELECTRICITY", "Symbol": "5110", "Owned_Qty": 3000, "Cost": 17.05, "Custodian": "Al Inma Capital"},
        {"Company": "SABIC", "Symbol": "2010", "Owned_Qty": 3000, "Cost": 71.97, "Custodian": "Al Inma Capital"},
        {"Company": "YANSAB", "Symbol": "2290", "Owned_Qty": 3877, "Cost": 34.34, "Custodian": "Al Inma Capital"},
        {"Company": "RETAL", "Symbol": "4322", "Owned_Qty": 4000, "Cost": 14.43, "Custodian": "Al Inma Capital"},
        {"Company": "JARIR", "Symbol": "4190", "Owned_Qty": 8573, "Cost": 13.57, "Custodian": "Al Inma Capital"},
        {"Company": "QACCO", "Symbol": "3040", "Owned_Qty": 972, "Cost": 62.51, "Custodian": "Al Inma Capital"},
        {"Company": "ALBILAD SAUDI GROWTH", "Symbol": "9408", "Owned_Qty": 1000, "Cost": 11.19, "Custodian": "Al Inma Capital"},
        {"Company": "A.OTHAIM MARKET", "Symbol": "4001", "Owned_Qty": 9500, "Cost": 10.75, "Custodian": "Al Inma Capital"},
        {"Company": "BINDAWOOD", "Symbol": "4161", "Owned_Qty": 4000, "Cost": 7.79, "Custodian": "Al Inma Capital"},
        {"Company": "ALAHLI REIT 1", "Symbol": "4338", "Owned_Qty": 3000, "Cost": 8.11, "Custodian": "Al Inma Capital"},
        {"Company": "BJAZ", "Symbol": "1020", "Owned_Qty": 10546, "Cost": 13.67, "Custodian": "Al Inma Capital"},
        {"Company": "EIC", "Symbol": "1303", "Owned_Qty": 2950, "Cost": 5.13, "Custodian": "BSF Capital"},
        {"Company": "TASNEE", "Symbol": "2060", "Owned_Qty": 5500, "Cost": 10.86, "Custodian": "BSF Capital"},
        {"Company": "SAVOLA GROUP", "Symbol": "2050", "Owned_Qty": 1061, "Cost": 62.74, "Custodian": "BSF Capital"},
        {"Company": "SAUDI ARAMCO", "Symbol": "2222", "Owned_Qty": 2550, "Cost": 26.61, "Custodian": "BSF Capital"},
        {"Company": "SAUDI DARB", "Symbol": "4130", "Owned_Qty": 2000, "Cost": 4.71, "Custodian": "BSF Capital"},
        {"Company": "ALYAMAMAH STEEL", "Symbol": "1304", "Owned_Qty": 1000, "Cost": 38.69, "Custodian": "BSF Capital"},
        {"Company": "SPIMACO", "Symbol": "2070", "Owned_Qty": 1500, "Cost": 31.99, "Custodian": "BSF Capital"},
        {"Company": "ALMARAI", "Symbol": "2280", "Owned_Qty": 646, "Cost": 0.00, "Custodian": "BSF Capital"},
        {"Company": "BATIC", "Symbol": "4110", "Owned_Qty": 6800, "Cost": 3.62, "Custodian": "BSF Capital"},
        {"Company": "DERAYAH", "Symbol": "4084", "Owned_Qty": 10, "Cost": 30.00, "Custodian": "BSF Capital"},
        {"Company": "MASAR", "Symbol": "4325", "Owned_Qty": 1015, "Cost": 23.00, "Custodian": "BSF Capital"},
        {"Company": "SISCO HOLDING", "Symbol": "2190", "Owned_Qty": 1000, "Cost": 28.83, "Custodian": "Al Rajhi Capital"},
        {"Company": "Riyadh Bank", "Symbol": "1010", "Owned_Qty": 3811, "Cost": 27.58, "Custodian": "Al Rajhi Capital"},
        {"Company": "ARAMCO", "Symbol": "2222", "Owned_Qty": 231, "Cost": 27.25, "Custodian": "Al Rajhi Capital"},
        {"Company": "Electrical Industries Co.", "Symbol": "1303", "Owned_Qty": 7500, "Cost": 7.13, "Custodian": "Al Rajhi Capital"},
        {"Company": "Al Ahli Bank", "Symbol": "1180", "Owned_Qty": 2000, "Cost": 33.43, "Custodian": "Al Rajhi Capital"},
        {"Company": "STC", "Symbol": "7010", "Owned_Qty": 3000, "Cost": 41.39, "Custodian": "Al Rajhi Capital"},
        {"Company": "Arabi Bank", "Symbol": "1080", "Owned_Qty": 3666, "Cost": 20.32, "Custodian": "Al Rajhi Capital"},
        {"Company": "Zain KSA", "Symbol": "7030", "Owned_Qty": 9400, "Cost": 11.41, "Custodian": "Al Rajhi Capital"},
        {"Company": "SAUDI DARB", "Symbol": "4130", "Owned_Qty": 6470, "Cost": 4.02, "Custodian": "Al Rajhi Capital"},
        {"Company": "ADES", "Symbol": "2382", "Owned_Qty": 1000, "Cost": 14.72, "Custodian": "Al Rajhi Capital"},
        {"Company": "CHEMICAL", "Symbol": "2230", "Owned_Qty": 1000, "Cost": 7.62, "Custodian": "Al Rajhi Capital"},
        {"Company": "Al Rajhi Bank", "Symbol": "1120", "Owned_Qty": 500, "Cost": 92.60, "Custodian": "Al Rajhi Capital"},
    ]
    
    # Create DataFrame
    df = pd.DataFrame(corrected_portfolio)
    
    # Create Excel file with formatting
    with pd.ExcelWriter('portfolio_corrected_costs.xlsx', engine='openpyxl') as writer:
        # Write corrected data
        df.to_excel(writer, sheet_name='Your Portfolio', index=False)
        
        # Write empty template
        empty_df = pd.DataFrame(columns=['Symbol', 'Company', 'Owned_Qty', 'Cost', 'Custodian'])
        empty_df.to_excel(writer, sheet_name='Empty Template', index=False)
        
        # Format the sheets
        workbook = writer.book
        
        # Format Your Portfolio sheet
        worksheet1 = writer.sheets['Your Portfolio']
        
        # Header formatting
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0d4f3c', end_color='0d4f3c', fill_type='solid')
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet1.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        worksheet1.column_dimensions['A'].width = 15  # Symbol
        worksheet1.column_dimensions['B'].width = 30  # Company
        worksheet1.column_dimensions['C'].width = 12  # Owned_Qty
        worksheet1.column_dimensions['D'].width = 12  # Cost
        worksheet1.column_dimensions['E'].width = 20  # Custodian
        
        # Format Empty Template sheet
        worksheet2 = writer.sheets['Empty Template']
        
        for col in range(1, len(empty_df.columns) + 1):
            cell = worksheet2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        worksheet2.column_dimensions['A'].width = 15
        worksheet2.column_dimensions['B'].width = 30
        worksheet2.column_dimensions['C'].width = 12
        worksheet2.column_dimensions['D'].width = 12
        worksheet2.column_dimensions['E'].width = 20
    
    print("✅ Created portfolio_corrected_costs.xlsx with:")
    print(f"   - Your Portfolio sheet: {len(df)} positions")
    print(f"   - Empty Template sheet for new entries")
    print(f"   - Total portfolio value: {(df['Owned_Qty'] * df['Cost']).sum():,.2f} SAR")
    
    # Show SISCO correction specifically
    sisco = df[df['Company'] == 'SISCO HOLDING'].iloc[0]
    print(f"\n✅ SISCO HOLDING corrected:")
    print(f"   Symbol: {sisco['Symbol']}")
    print(f"   Quantity: {sisco['Owned_Qty']}")
    print(f"   Corrected Cost: {sisco['Cost']:.2f} SAR (was 34.5)")
    print(f"   Total Value: {sisco['Owned_Qty'] * sisco['Cost']:,.2f} SAR")

if __name__ == "__main__":
    create_corrected_portfolio_template()
